import shutil
import subprocess
import tempfile
from pathlib import Path


def extract(filepath: Path) -> list[tuple[int, str]]:
    """Return list of (page_number, text) tuples. Pages are 1-indexed."""
    ext = filepath.suffix.lower()
    if ext == ".pdf":
        return _extract_pdf(filepath)
    elif ext == ".docx":
        return _extract_docx(filepath)
    elif ext == ".doc":
        return _extract_doc(filepath)
    elif ext == ".xlsx":
        return _extract_xlsx(filepath)
    elif ext == ".xls":
        return _extract_xls(filepath)
    elif ext == ".txt":
        return _extract_txt(filepath)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def _extract_pdf(filepath: Path) -> list[tuple[int, str]]:
    import fitz  # pymupdf

    pages = []
    with fitz.open(str(filepath)) as doc:
        for i, page in enumerate(doc, start=1):
            text = page.get_text()
            if text.strip():
                pages.append((i, text))
    return pages


# ---------------------------------------------------------------------------
# Word — DOCX
# ---------------------------------------------------------------------------

def _extract_docx(filepath: Path) -> list[tuple[int, str]]:
    from docx import Document

    doc = Document(str(filepath))
    pages: list[tuple[int, str]] = []
    current_page = 1
    current_lines: list[str] = []

    for para in doc.paragraphs:
        if para.contains_page_break:
            if current_lines:
                pages.append((current_page, "\n".join(current_lines)))
                current_lines = []
            current_page += 1
        current_lines.append(para.text)

    if current_lines:
        pages.append((current_page, "\n".join(current_lines)))

    return pages


# ---------------------------------------------------------------------------
# Word — DOC  (requires LibreOffice)
# ---------------------------------------------------------------------------

def _find_soffice() -> str | None:
    """Return path to the LibreOffice soffice executable, or None."""
    for name in ("libreoffice", "soffice"):
        path = shutil.which(name)
        if path:
            return path
    # Windows common install locations
    import os
    for root in (
        os.environ.get("PROGRAMFILES", ""),
        os.environ.get("PROGRAMFILES(X86)", ""),
        r"C:\Program Files",
        r"C:\Program Files (x86)",
    ):
        candidate = Path(root) / "LibreOffice" / "program" / "soffice.exe"
        if candidate.exists():
            return str(candidate)
    return None


def _extract_doc(filepath: Path) -> list[tuple[int, str]]:
    soffice = _find_soffice()
    if not soffice:
        raise RuntimeError(
            "LibreOffice not found. Install it to index .doc files "
            "(https://www.libreoffice.org)."
        )
    with tempfile.TemporaryDirectory() as tmpdir:
        subprocess.run(
            [soffice, "--headless", "--convert-to", "txt:Text",
             "--outdir", tmpdir, str(filepath)],
            capture_output=True, timeout=120,
        )
        txt_file = Path(tmpdir) / (filepath.stem + ".txt")
        if not txt_file.exists():
            return []
        text = txt_file.read_text(encoding="utf-8", errors="replace")
    return [(1, text)] if text.strip() else []


# ---------------------------------------------------------------------------
# Excel — XLSX
# ---------------------------------------------------------------------------

def _extract_xlsx(filepath: Path) -> list[tuple[int, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(str(filepath), read_only=True, data_only=True)
    pages = []
    for page_num, sheet in enumerate(wb.worksheets, start=1):
        lines = []
        for row in sheet.iter_rows(values_only=True):
            parts = [str(cell) for cell in row if cell is not None and str(cell).strip()]
            if parts:
                lines.append("  ".join(parts))
        if lines:
            pages.append((page_num, "\n".join(lines)))
    wb.close()
    return pages


# ---------------------------------------------------------------------------
# Excel — XLS
# ---------------------------------------------------------------------------

def _extract_xls(filepath: Path) -> list[tuple[int, str]]:
    import xlrd

    wb = xlrd.open_workbook(str(filepath))
    pages = []
    for page_num, sheet in enumerate(wb.sheets(), start=1):
        lines = []
        for row_idx in range(sheet.nrows):
            parts = [
                str(sheet.cell_value(row_idx, col))
                for col in range(sheet.ncols)
                if str(sheet.cell_value(row_idx, col)).strip()
            ]
            if parts:
                lines.append("  ".join(parts))
        if lines:
            pages.append((page_num, "\n".join(lines)))
    return pages


# ---------------------------------------------------------------------------
# Plain text
# ---------------------------------------------------------------------------

def _extract_txt(filepath: Path) -> list[tuple[int, str]]:
    text = filepath.read_text(encoding="utf-8", errors="replace")
    return [(1, text)] if text.strip() else []
